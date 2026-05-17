from __future__ import annotations

import re
from datetime import datetime
from io import BytesIO
from typing import Any


def split_product_blocks(text: str) -> list[str]:
    clean = text.strip()
    if not clean:
        return []
    blocks = [block.strip() for block in re.split(r"\n\s*\n", clean) if block.strip()]
    return blocks or [clean]


def infer_product_name(block: str, index: int) -> str:
    for line in block.splitlines():
        line = line.strip()
        if not line:
            continue
        if any(label in line for label in ["商品名", "品名", "Product"]):
            return re.sub(r"^(商品名|品名|Product)\s*[:：]\s*", "", line).strip()
        return line[:40]
    return f"商品{index}"


def build_product_rows(product_text: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for index, block in enumerate(split_product_blocks(product_text), start=1):
        product_name = infer_product_name(block, index)
        compact = " ".join(line.strip() for line in block.splitlines() if line.strip())
        rows.append(
            {
                "商品名": product_name,
                "型番": "要確認",
                "商品説明文": compact[:240] if compact else "要確認",
                "箇条書き1": "商品の主な特徴を確認してください。",
                "箇条書き2": "素材・サイズ・同梱物などの不足情報を確認してください。",
                "箇条書き3": "利用シーンや対象ユーザーを明確にしてください。",
                "検索キーワード": "要確認",
                "要確認項目": "型番、サイズ、素材、JAN、禁止表現の確認",
            }
        )
    return rows


def create_product_workbook(
    template_bytes: bytes | None,
    product_text: str,
    memo: str,
) -> tuple[bytes, dict[str, Any]]:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill
    except Exception as exc:
        raise RuntimeError("openpyxl が未インストールです。requirements.txt を入れてください。") from exc

    rows = build_product_rows(product_text)
    if template_bytes:
        workbook = load_workbook(BytesIO(template_bytes))
    else:
        workbook = Workbook()
        workbook.active.title = "template"

    sheet_name = "AI整理結果"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    sheet = workbook.create_sheet(sheet_name, 0)

    headers = [
        "商品名",
        "型番",
        "商品説明文",
        "箇条書き1",
        "箇条書き2",
        "箇条書き3",
        "検索キーワード",
        "要確認項目",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1D4ED8")

    for row in rows:
        sheet.append([row.get(header, "") for header in headers])

    sheet.append([])
    sheet.append(["補足メモ", memo or "なし"])
    sheet.append(["作成日時", datetime.now().strftime("%Y-%m-%d %H:%M")])

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 48)

    output = BytesIO()
    workbook.save(output)
    summary = {
        "product_count": len(rows),
        "needs_review": sum(1 for row in rows if "要確認" in " ".join(row.values())),
        "sheet_name": sheet_name,
    }
    return output.getvalue(), summary

