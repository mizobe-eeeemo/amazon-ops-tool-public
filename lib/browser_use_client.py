from __future__ import annotations

import csv
import gzip
import io
import json
import os
import re
import time
import zipfile
from dataclasses import dataclass
from datetime import datetime, timedelta
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen
from zoneinfo import ZoneInfo

from lib.auth import get_secret


@dataclass(frozen=True)
class BrowserUseConfig:
    api_key_configured: bool
    seller_central_profile_id: str | None
    seller_central_profile_id_b: str | None
    default_model: str
    max_cost_usd: str
    proxy_country_code: str | None
    poll_timeout_seconds: int


@dataclass(frozen=True)
class BrowserUseRunResult:
    attempted: bool
    status: str
    summary: str
    session_id: str | None = None
    live_url: str | None = None
    last_step_summary: str | None = None
    output: Any | None = None
    is_success: bool | None = None
    total_cost_usd: str | None = None
    error: str | None = None
    screenshot_url: str | None = None

    @property
    def finished(self) -> bool:
        return self.status in {"stopped", "timed_out", "error", "unavailable", "failed"}

    def prompt_context(self) -> str:
        payload = {
            "attempted": self.attempted,
            "status": self.status,
            "summary": self.summary,
            "session_id": self.session_id,
            "last_step_summary": self.last_step_summary,
            "output": self.output,
            "is_success": self.is_success,
            "total_cost_usd": self.total_cost_usd,
            "error": self.error,
            "screenshot": "取得あり" if self.screenshot_url else "なし",
        }
        return json.dumps(payload, ensure_ascii=False, indent=2)


class BrowserUseError(RuntimeError):
    pass


def _secret_or_env(name: str, default: str | None = None) -> str | None:
    value = get_secret(name) or os.getenv(name) or default
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def get_browser_use_config() -> BrowserUseConfig:
    api_key = _secret_or_env("BROWSER_USE_API_KEY")
    profile_id = _secret_or_env("BROWSER_USE_PROFILE_ID_SELLER_CENTRAL")
    profile_id_b = _secret_or_env("BROWSER_USE_PROFILE_ID_SELLER_CENTRAL_B")
    default_model = _secret_or_env("BROWSER_USE_MODEL", "bu-mini") or "bu-mini"
    max_cost_usd = _secret_or_env("BROWSER_USE_MAX_COST_USD", "0.25") or "0.25"
    proxy_country_code = _secret_or_env("BROWSER_USE_PROXY_COUNTRY_CODE", "jp")
    if proxy_country_code and proxy_country_code.lower() in {"none", "off", "false"}:
        proxy_country_code = None
    timeout_raw = _secret_or_env("BROWSER_USE_POLL_TIMEOUT_SECONDS", "75") or "75"
    try:
        poll_timeout_seconds = max(10, min(180, int(timeout_raw)))
    except ValueError:
        poll_timeout_seconds = 90
    return BrowserUseConfig(
        api_key_configured=bool(api_key),
        seller_central_profile_id=profile_id,
        seller_central_profile_id_b=profile_id_b,
        default_model=default_model,
        max_cost_usd=max_cost_usd,
        proxy_country_code=proxy_country_code,
        poll_timeout_seconds=poll_timeout_seconds,
    )


def browser_use_available() -> bool:
    return get_browser_use_config().api_key_configured


def seller_central_profile_available() -> bool:
    config = get_browser_use_config()
    return config.api_key_configured and bool(config.seller_central_profile_id)


def seller_central_profile_b_available() -> bool:
    config = get_browser_use_config()
    return config.api_key_configured and bool(config.seller_central_profile_id_b)


def profile_available_for_account(account_key: str | None) -> bool:
    config = get_browser_use_config()
    if not config.api_key_configured:
        return False
    if account_key == "A":
        return bool(config.seller_central_profile_id)
    if account_key == "B":
        return bool(config.seller_central_profile_id_b)
    return False


def seller_central_profile_id_for_account(account_key: str | None) -> str | None:
    config = get_browser_use_config()
    if account_key == "A":
        return config.seller_central_profile_id
    if account_key == "B":
        return config.seller_central_profile_id_b
    return None


def _get_api_key() -> str | None:
    return _secret_or_env("BROWSER_USE_API_KEY")


def _normalize_v3_model(model: str) -> str:
    mapping = {
        "browser-use-2.0": "bu-mini",
        "browser-use-llm": "bu-mini",
        "claude-sonnet-4-6": "claude-sonnet-4.6",
        "claude-sonnet-4-5-20250929": "claude-sonnet-4.6",
        "claude-opus-4-7": "claude-opus-4.7",
    }
    return mapping.get(model, model)


def _last_completed_month_range(today: datetime) -> tuple[str, str]:
    first_this_month = today.date().replace(day=1)
    last_month_end = first_this_month - timedelta(days=1)
    last_month_start = last_month_end.replace(day=1)
    return last_month_start.isoformat(), last_month_end.isoformat()


def _date_range_instruction(question: str, today: datetime) -> str:
    if "先月" in question:
        start_date, end_date = _last_completed_month_range(today)
        return (
            f'The user asked for "先月". Use the absolute date range {start_date} through {end_date} '
            'in Japan time. Do not search the page for the word "今日". Use the date range control, '
            'choose a "Last month" / "先月" preset if it exists, or enter/select these exact start and end dates.'
        )
    return (
        "Use the date range implied by the user question. If the period is ambiguous, report which current "
        "date range is visible instead of repeatedly trying to change the filter."
    )


def _is_report_pickup_question(question: str) -> bool:
    keywords = [
        "レポート一覧",
        "既存レポート",
        "作成済みレポート",
        "直近作成済み",
        "完了していれば",
        "保留中",
        "処理中",
        "重複作成しない",
        "ダウンロード未完了",
        "該当行",
    ]
    return any(keyword in question for keyword in keywords)


def is_amazon_ads_report_pickup_question(question: str) -> bool:
    return _is_report_pickup_question(question)


def _report_workflow_instruction(question: str) -> str:
    if _is_report_pickup_question(question):
        return (
            "Report pickup mode: a matching report may already exist from a previous run. Do not create a duplicate report. "
            "Open the reports list, use browser reload once, reopen the One-time / 1回限り tab if needed, and find the matching report. "
            "If it is ready, download it. If it is still pending after one more reload check, return status blocked with blocked_by "
            '"report_processing_not_ready". If you cannot confidently identify the matching row after one reload, return status blocked '
            'with blocked_by "matching_report_not_found". Do not run broad JavaScript/page-wide searches or keep scanning the whole page. '
            "Keep this run focused on downloading an existing report."
        )
    return (
        "Report creation mode: create the requested report if no matching report exists, then check the reports list. "
        "If the matching report is pending, refresh the browser page at most twice. If it is still pending, return blocked instead of waiting until the session cost limit."
    )


def _browser_use_request(method: str, path: str, payload: dict[str, Any] | None = None) -> dict[str, Any]:
    api_key = _get_api_key()
    if not api_key:
        raise BrowserUseError("BROWSER_USE_API_KEY is not configured.")

    url = f"https://api.browser-use.com{path}"
    body = None
    headers = {"X-Browser-Use-API-Key": api_key}
    if payload is not None:
        body = json.dumps(payload).encode("utf-8")
        headers["Content-Type"] = "application/json"

    request = Request(url, data=body, headers=headers, method=method)
    try:
        with urlopen(request, timeout=60) as response:
            raw = response.read().decode("utf-8")
    except HTTPError as exc:
        error_body = exc.read().decode("utf-8", errors="replace")
        raise BrowserUseError(f"browser-use API error {exc.code}: {error_body[:500]}") from exc
    except URLError as exc:
        raise BrowserUseError(f"browser-use API connection error: {exc.reason}") from exc

    if not raw:
        return {}
    try:
        return json.loads(raw)
    except json.JSONDecodeError as exc:
        raise BrowserUseError("browser-use API returned invalid JSON.") from exc


def _browser_use_downloads(session_id: str) -> dict[str, Any]:
    return _browser_use_request("GET", f"/api/v3/browsers/{session_id}/downloads?includeUrls=true")


def stop_browser_use_session(session_id: str) -> BrowserUseRunResult:
    cleaned_session_id = session_id.strip()
    if not cleaned_session_id:
        return BrowserUseRunResult(
            attempted=False,
            status="unavailable",
            summary="停止するbrowser-useセッションIDを入力してください。",
        )
    try:
        session = _browser_use_request(
            "POST",
            f"/api/v3/sessions/{cleaned_session_id}/stop",
            {"strategy": "session"},
        )
        return _result_from_session(session, summary="browser-useセッションを停止しました。")
    except BrowserUseError as exc:
        return BrowserUseRunResult(
            attempted=True,
            status="error",
            summary="browser-useセッション停止に失敗しました。",
            session_id=cleaned_session_id,
            error=str(exc),
        )


def get_browser_use_session(session_id: str) -> BrowserUseRunResult:
    cleaned_session_id = session_id.strip()
    if not cleaned_session_id:
        return BrowserUseRunResult(
            attempted=False,
            status="unavailable",
            summary="確認するbrowser-useセッションIDを入力してください。",
        )
    try:
        session = _browser_use_request("GET", f"/api/v3/sessions/{cleaned_session_id}")
        return _result_from_session(session, summary="browser-useセッション状態を取得しました。")
    except BrowserUseError as exc:
        return BrowserUseRunResult(
            attempted=True,
            status="error",
            summary="browser-useセッション状態の取得に失敗しました。",
            session_id=cleaned_session_id,
            error=str(exc),
        )


def _output_status(output: Any) -> str:
    if isinstance(output, dict):
        return str(output.get("status") or "")
    return ""


def _output_summary(output: Any) -> str:
    if isinstance(output, dict):
        return str(output.get("summary") or "")
    return ""


def _output_is_success(output: Any) -> bool:
    return _output_status(output).lower() == "success"


def _output_has_result(output: Any) -> bool:
    return _output_status(output).lower() in {"success", "partial", "blocked"}


def _safe_download_path(path: Any) -> str | None:
    if path is None:
        return None
    text = str(path).split("?", 1)[0]
    if "://" in text:
        text = text.rstrip("/").rsplit("/", 1)[-1]
    return text or None


def _safe_file_info(file_info: dict[str, Any]) -> dict[str, Any]:
    return {
        "path": _safe_download_path(file_info.get("path")),
        "size": file_info.get("size"),
        "lastModified": file_info.get("lastModified"),
    }


def _safe_error_text(error: Any) -> str:
    text = str(error)
    text = re.sub(r"https?://[^\s\"'<>]+", "[URL omitted]", text)
    return text[:500]


def _metric_number(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = (
            value.replace(",", "")
            .replace("¥", "")
            .replace("￥", "")
            .replace("%", "")
            .strip()
        )
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _parse_metric_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text in {"-", "--", "—", "N/A", "n/a"}:
        return None
    negative = text.startswith("(") and text.endswith(")")
    cleaned = (
        text.replace(",", "")
        .replace("¥", "")
        .replace("￥", "")
        .replace("$", "")
        .replace("%", "")
        .replace("円", "")
        .replace("クリック", "")
        .strip("() ")
    )
    cleaned = re.sub(r"[^0-9.\-]", "", cleaned)
    if not cleaned or cleaned in {"-", "."}:
        return None
    try:
        number = float(cleaned)
    except ValueError:
        return None
    return -number if negative else number


def _normalize_header(value: Any) -> str:
    return re.sub(r"[\s_\-()./%（）・/]+", "", str(value or "").strip().lower())


REPORT_COLUMN_ALIASES = {
    "ad_spend": [
        "spend",
        "cost",
        "totalcost",
        "totalspend",
        "広告費",
        "費用",
        "支出",
        "消化金額",
        "コスト",
    ],
    "ad_sales": [
        "sales",
        "adsales",
        "totalsales",
        "sales1d",
        "sales7d",
        "sales14d",
        "sales30d",
        "attributedsales",
        "広告売上",
        "売上",
        "売上高",
    ],
    "clicks": ["clicks", "クリック", "クリック数"],
    "impressions": ["impressions", "インプレッション", "表示回数"],
    "acos": ["acos", "acosclicks7d", "acosclicks14d", "広告費売上高比率"],
    "cpc": ["cpc", "costperclick", "クリック単価"],
    "ctr": ["ctr", "clickthroughrate", "クリック率"],
    "orders": ["orders", "purchases", "purchases7d", "purchases14d", "注文", "注文数"],
}


def _looks_like_report_header(row: list[Any]) -> bool:
    normalized = [_normalize_header(cell) for cell in row]
    matches = 0
    for aliases in REPORT_COLUMN_ALIASES.values():
        if any(any(alias in header for header in normalized) for alias in aliases):
            matches += 1
    return matches >= 2


def _rows_from_grid(rows: list[list[Any]]) -> list[dict[str, Any]]:
    header_index = None
    for index, row in enumerate(rows[:30]):
        if _looks_like_report_header(row):
            header_index = index
            break
    if header_index is None:
        return []

    headers = [str(cell or "").strip() for cell in rows[header_index]]
    records: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        if not any(str(cell or "").strip() for cell in row):
            continue
        record = {
            headers[index]: row[index] if index < len(row) else ""
            for index in range(len(headers))
            if headers[index]
        }
        if record:
            records.append(record)
    return records


def _rows_from_csv_bytes(data: bytes, filename: str = "") -> list[dict[str, Any]]:
    text = None
    for encoding in ("utf-8-sig", "utf-8", "cp932", "shift_jis"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        return []
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        first_line = text.splitlines()[0] if text.splitlines() else ""
        delimiter = "\t" if filename.lower().endswith(".tsv") or first_line.count("\t") > first_line.count(",") else ","
        rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    else:
        rows = list(csv.reader(io.StringIO(text), dialect))
    return _rows_from_grid(rows)


def _rows_from_xlsx_bytes(data: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return []
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    for worksheet in workbook.worksheets:
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        records = _rows_from_grid(rows)
        if records:
            return records
    return []


def _rows_from_report_bytes(data: bytes, filename: str) -> list[dict[str, Any]]:
    lower_name = filename.lower()
    if lower_name.endswith(".gz"):
        try:
            return _rows_from_report_bytes(gzip.decompress(data), lower_name[:-3])
        except OSError:
            return []
    if lower_name.endswith(".zip"):
        try:
            with zipfile.ZipFile(io.BytesIO(data)) as archive:
                for member in archive.namelist():
                    if member.lower().endswith((".csv", ".tsv", ".xlsx", ".gz")):
                        with archive.open(member) as file:
                            rows = _rows_from_report_bytes(file.read(), member)
                        if rows:
                            return rows
        except zipfile.BadZipFile:
            return []
        return []
    if lower_name.endswith((".xlsx", ".xlsm")):
        return _rows_from_xlsx_bytes(data)
    return _rows_from_csv_bytes(data, lower_name)


def _find_report_value(row: dict[str, Any], metric: str) -> Any:
    aliases = REPORT_COLUMN_ALIASES[metric]
    for key, value in row.items():
        normalized = _normalize_header(key)
        if any(alias in normalized for alias in aliases):
            return value
    return None


def _aggregate_report_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    totals = {
        "ad_spend": 0.0,
        "ad_sales": 0.0,
        "clicks": 0.0,
        "impressions": 0.0,
        "orders": 0.0,
    }
    seen = {key: False for key in totals}
    acos_values: list[float] = []
    cpc_values: list[float] = []
    ctr_values: list[float] = []

    for row in rows:
        for metric in totals:
            value = _parse_metric_number(_find_report_value(row, metric))
            if value is not None:
                totals[metric] += value
                seen[metric] = True
        for metric, values in (("acos", acos_values), ("cpc", cpc_values), ("ctr", ctr_values)):
            value = _parse_metric_number(_find_report_value(row, metric))
            if value is not None:
                values.append(value)

    metrics: dict[str, Any] = {
        "ad_spend": round(totals["ad_spend"], 2) if seen["ad_spend"] else None,
        "ad_sales": round(totals["ad_sales"], 2) if seen["ad_sales"] else None,
        "clicks": int(round(totals["clicks"])) if seen["clicks"] else None,
        "impressions": int(round(totals["impressions"])) if seen["impressions"] else None,
        "orders": int(round(totals["orders"])) if seen["orders"] else None,
        "acos": None,
        "cpc": None,
        "ctr": None,
    }
    if metrics["ad_spend"] is not None and metrics["ad_sales"]:
        metrics["acos"] = round(float(metrics["ad_spend"]) / float(metrics["ad_sales"]) * 100, 2)
    elif acos_values:
        metrics["acos"] = round(sum(acos_values) / len(acos_values), 2)
    if metrics["ad_spend"] is not None and metrics["clicks"]:
        metrics["cpc"] = round(float(metrics["ad_spend"]) / float(metrics["clicks"]), 2)
    elif cpc_values:
        metrics["cpc"] = round(sum(cpc_values) / len(cpc_values), 2)
    if metrics["clicks"] is not None and metrics["impressions"]:
        metrics["ctr"] = round(float(metrics["clicks"]) / float(metrics["impressions"]) * 100, 2)
    elif ctr_values:
        metrics["ctr"] = round(sum(ctr_values) / len(ctr_values), 2)
    return metrics


def _download_report_file(file_info: dict[str, Any]) -> tuple[dict[str, Any] | None, str | None]:
    url = file_info.get("url")
    filename = str(file_info.get("path") or "report.csv")
    if not url:
        return None, "ダウンロードURLが取得できませんでした。"
    try:
        with urlopen(Request(str(url), headers={"User-Agent": "amazon-ops-tool/1.0"}), timeout=90) as response:
            data = response.read()
    except Exception as exc:
        return None, f"レポートファイルの取得に失敗しました: {_safe_error_text(exc)}"
    rows = _rows_from_report_bytes(data, filename)
    if not rows:
        return None, "レポートファイル内の表ヘッダーを認識できませんでした。"
    return {
        "file": _safe_file_info(file_info),
        "row_count": len(rows),
        "metrics": _aggregate_report_rows(rows),
    }, None


def _extract_downloaded_report_metrics(session_id: str) -> tuple[dict[str, Any] | None, list[str]]:
    warnings: list[str] = []
    try:
        downloads = _browser_use_downloads(session_id)
    except BrowserUseError as exc:
        return None, [f"browser-useダウンロード一覧の取得に失敗しました: {_safe_error_text(exc)}"]

    files = downloads.get("files") or []
    report_files = [
        file_info
        for file_info in files
        if str(file_info.get("path") or "").lower().endswith((".csv", ".tsv", ".xlsx", ".zip", ".gz"))
    ]
    report_files.sort(key=lambda item: str(item.get("lastModified") or ""), reverse=True)
    if not report_files:
        return {
            "downloaded_files": [_safe_file_info(file_info) for file_info in files],
            "metrics": {},
        }, ["ダウンロード済みレポートファイルが見つかりませんでした。"]

    for file_info in report_files:
        parsed, warning = _download_report_file(file_info)
        if parsed:
            parsed["downloaded_files"] = [_safe_file_info(item) for item in files]
            return parsed, warnings
        if warning:
            warnings.append(f"{file_info.get('path')}: {warning}")
    return {
        "downloaded_files": [_safe_file_info(file_info) for file_info in files],
        "metrics": {},
    }, warnings


def check_browser_use_downloads(session_id: str) -> BrowserUseRunResult:
    cleaned_session_id = session_id.strip()
    if not cleaned_session_id:
        return BrowserUseRunResult(
            attempted=False,
            status="unavailable",
            summary="確認するbrowser-useセッションIDを入力してください。",
        )

    parsed, warnings = _extract_downloaded_report_metrics(cleaned_session_id)
    if parsed is None:
        return BrowserUseRunResult(
            attempted=True,
            status="error",
            summary="browser-useダウンロード一覧の確認に失敗しました。",
            session_id=cleaned_session_id,
            error="\n".join(warnings),
            output={
                "status": "error",
                "summary": "browser-useダウンロード一覧の確認に失敗しました。",
                "source": "browser_use_downloads",
                "metrics": {},
                "downloaded_files": [],
                "notes": warnings,
            },
        )

    metrics = parsed.get("metrics") or {}
    downloaded_files = parsed.get("downloaded_files", [])
    has_exact_metrics = any(metrics.get(key) is not None for key in ("ad_spend", "ad_sales", "clicks"))
    if has_exact_metrics:
        output = {
            "status": "success",
            "summary": "前回セッションの広告レポートファイルから実績値を取得しました。",
            "source": "downloaded_ad_report",
            "metrics": metrics,
            "downloaded_report": parsed.get("file"),
            "downloaded_files": downloaded_files,
            "notes": warnings,
        }
        return BrowserUseRunResult(
            attempted=True,
            status="download_checked",
            summary="前回セッションの広告レポートファイルから実績値を取得しました。",
            session_id=cleaned_session_id,
            output=output,
            is_success=True,
        )

    blocked_by = "downloaded_report_parse_failed" if downloaded_files and warnings else "downloaded_report_not_found"
    output = {
        "status": "blocked",
        "summary": "前回セッション内に解析可能な広告レポートファイルはまだ見つかりませんでした。",
        "source": "browser_use_downloads",
        "metrics": metrics,
        "downloaded_files": downloaded_files,
        "notes": warnings,
        "blocked_by": blocked_by,
    }
    return BrowserUseRunResult(
        attempted=True,
        status="download_checked",
        summary="前回セッション内に解析可能な広告レポートファイルはまだ見つかりませんでした。",
        session_id=cleaned_session_id,
        output=output,
        is_success=False,
    )


def _merge_report_downloads(result: BrowserUseRunResult) -> BrowserUseRunResult:
    if not result.session_id:
        return result
    parsed, warnings = _extract_downloaded_report_metrics(result.session_id)
    if parsed is None:
        return result

    output = result.output if isinstance(result.output, dict) else {}
    output = dict(output)
    existing_notes = output.get("notes")
    notes = existing_notes if isinstance(existing_notes, list) else []
    notes.extend(warnings)
    report_metrics = parsed.get("metrics") or {}
    has_exact_metrics = any(report_metrics.get(key) is not None for key in ("ad_spend", "ad_sales", "clicks"))
    if has_exact_metrics:
        output.update(
            {
                "status": "success",
                "summary": "広告レポートファイルから実績値を取得しました。",
                "source": "downloaded_ad_report",
                "metrics": report_metrics,
                "downloaded_report": parsed.get("file"),
                "downloaded_files": parsed.get("downloaded_files", []),
                "notes": notes,
            }
        )
        return BrowserUseRunResult(
            attempted=result.attempted,
            status=result.status,
            summary="広告レポートファイルから実績値を取得しました。",
            session_id=result.session_id,
            live_url=result.live_url,
            last_step_summary=result.last_step_summary,
            output=output,
            is_success=True,
            total_cost_usd=result.total_cost_usd,
            error=result.error,
            screenshot_url=result.screenshot_url,
        )

    output["downloaded_files"] = parsed.get("downloaded_files", [])
    output["notes"] = notes
    return BrowserUseRunResult(
        attempted=result.attempted,
        status=result.status,
        summary=result.summary,
        session_id=result.session_id,
        live_url=result.live_url,
        last_step_summary=result.last_step_summary,
        output=output,
        is_success=result.is_success,
        total_cost_usd=result.total_cost_usd,
        error=result.error,
        screenshot_url=result.screenshot_url,
    )


def _enrich_estimated_metrics(output: Any) -> Any:
    if not isinstance(output, dict):
        return output
    if not output.get("allow_estimates"):
        return output
    metrics = output.get("metrics")
    if not isinstance(metrics, dict):
        return output

    clicks = _metric_number(metrics.get("clicks"))
    cpc = _metric_number(metrics.get("cpc"))
    ad_sales = _metric_number(metrics.get("ad_sales"))
    ad_spend = _metric_number(metrics.get("ad_spend"))
    estimated: dict[str, float] = {}

    if ad_spend is None and clicks is not None and cpc is not None:
        estimated_ad_spend = round(clicks * cpc)
        estimated["ad_spend_from_clicks_cpc"] = estimated_ad_spend
        if ad_sales and ad_sales > 0:
            estimated["acos_from_estimated_spend"] = round(estimated_ad_spend / ad_sales * 100, 2)

    if estimated:
        output["estimated_metrics"] = estimated
        notes = output.get("notes")
        if not isinstance(notes, list):
            notes = []
        notes.append(
            "広告費は画面上で未取得です。estimated_metricsはクリック数×CPCから計算した概算で、実広告費ではありません。"
        )
        output["notes"] = notes
    return output


SELLER_CENTRAL_OUTPUT_SCHEMA: dict[str, Any] = {
    "type": "object",
    "properties": {
        "status": {
            "type": "string",
            "description": "success, partial, needs_user_login, needs_confirmation, unavailable, or error",
        },
        "summary": {"type": "string"},
        "shop_name": {"type": ["string", "null"]},
        "period": {"type": ["string", "null"]},
        "currency": {"type": ["string", "null"]},
        "current_url": {"type": ["string", "null"]},
        "visible_screen": {"type": ["string", "null"]},
        "blocked_by": {"type": ["string", "null"]},
        "source": {"type": ["string", "null"]},
        "report_scope": {"type": ["string", "null"]},
        "downloaded_report": {"type": ["object", "null"], "additionalProperties": True},
        "downloaded_files": {"type": "array", "items": {"type": "object", "additionalProperties": True}},
        "metrics": {
            "type": "object",
            "properties": {
                "ad_spend": {"type": ["number", "string", "null"]},
                "ad_sales": {"type": ["number", "string", "null"]},
                "acos": {"type": ["number", "string", "null"]},
                "roas": {"type": ["number", "string", "null"]},
                "impressions": {"type": ["number", "string", "null"]},
                "clicks": {"type": ["number", "string", "null"]},
                "ctr": {"type": ["number", "string", "null"]},
                "cpc": {"type": ["number", "string", "null"]},
                "orders": {"type": ["number", "string", "null"]},
                "cvr": {"type": ["number", "string", "null"]},
            },
            "additionalProperties": True,
        },
        "notes": {"type": "array", "items": {"type": "string"}},
    },
    "required": ["status", "summary", "shop_name", "period", "metrics", "notes"],
    "additionalProperties": True,
}

SELLER_CENTRAL_ACCESS_CHECK_SCHEMA: dict[str, Any] = {
    "type": "object",
    "properties": {
        "status": {
            "type": "string",
            "description": "success, needs_user_login, needs_2fa, needs_confirmation, blocked, or error",
        },
        "summary": {"type": "string"},
        "shop_name": {"type": ["string", "null"]},
        "current_url": {"type": ["string", "null"]},
        "page_title": {"type": ["string", "null"]},
        "reached_seller_central": {"type": "boolean"},
        "reached_target_shop": {"type": "boolean"},
        "reached_ad_screen": {"type": "boolean"},
        "blocked_by": {"type": ["string", "null"]},
        "visible_screen": {"type": "string"},
        "notes": {"type": "array", "items": {"type": "string"}},
    },
    "required": [
        "status",
        "summary",
        "shop_name",
        "current_url",
        "page_title",
        "reached_seller_central",
        "reached_target_shop",
        "reached_ad_screen",
        "blocked_by",
        "visible_screen",
        "notes",
    ],
    "additionalProperties": True,
}


def _build_seller_central_task(client: dict[str, Any], question: str) -> str:
    now = datetime.now(ZoneInfo("Asia/Tokyo"))
    today = now.date().isoformat()
    date_range_instruction = _date_range_instruction(question, now)
    report_workflow_instruction = _report_workflow_instruction(question)
    company_name = client.get("name") or ""
    shop_name = client.get("shop_name") or ""
    marketplace = client.get("marketplace") or "Amazon.co.jp"
    return f"""
You are controlling an already-authenticated browser profile for Amazon Seller Central Japan.

Safety rules:
- You are allowed to create and download Amazon Ads performance reports without asking the user again. The user has already approved report creation for this workflow.
- Do not type, request, reveal, or store passwords, 2FA codes, API keys, recovery codes, or other secrets.
- If Amazon asks for login, password, passkey, CAPTCHA, or 2FA, stop immediately and return status "needs_user_login".
- Do not change bids, budgets, campaigns, listings, account settings, billing settings, or payments.
- Do not click any button that confirms a purchase, subscription, payment, campaign launch, paid upgrade, or billing change.
- If a paid confirmation or billing confirmation appears, stop immediately and return status "needs_confirmation".

Client:
- Company name: {company_name}
- Seller Central shop/store name to select: {shop_name}
- Marketplace: {marketplace}
- Today in Japan: {today}
- User question: {question}
- Date instruction: {date_range_instruction}
- Report workflow instruction: {report_workflow_instruction}

Task:
1. Open the Amazon Ads console directly: https://advertising.amazon.co.jp/
2. Confirm that the active advertiser/store/account matches "{shop_name}" or visibly shows "{shop_name}". If a selector is visible and an exact or very close match exists, choose it.
3. Navigate to Sponsored ads reports / 広告レポート / レポート. Use direct report pages if available from the UI; otherwise use the visible Reports navigation.
4. Follow the Report workflow instruction above. Before creating anything new, inspect the reports list for a recently created report that matches the requested date range and campaign/Sponsored Products scope. If it is complete or has a download icon/button, download it immediately.
5. If the matching report is visible but still processing, use the refresh behavior from the Report workflow instruction. Do not wait until the session cost limit. Do not create a duplicate while a matching report is still processing.
6. If no matching report exists and the Report workflow instruction is creation mode, create a downloadable campaign performance report for the requested date range. The report creation itself is approved. Do not stop merely because a "Create report" / "レポートを作成" button is shown.
7. Prefer an all-sponsored-ads campaign report if the UI offers it. If the UI requires separate ad products, create Sponsored Products campaign report first. Include report_scope in the result so we know whether it is all ads or Sponsored Products only.
8. Use the date instruction above. For April 2026, use 2026-04-01 through 2026-04-30. Do not search the campaign manager for "今日".
9. Choose CSV or Excel/XLSX if format is selectable. Prefer summary/campaign level. Include columns for spend/cost, sales, ACOS, clicks, impressions, CPC, and CTR when selectable.
10. After creating a report, return to the reports list and wait for it to finish. Every 45 seconds, refresh the browser page, reopen the One-time / 1回限り tab if needed, and re-check the matching row. Repeat up to 4 times. Download the report as soon as a download button/icon appears.
11. If the report is still processing after those checks, return status "blocked" with blocked_by "report_processing_not_ready", current_url, visible_screen, and a note that the report was created but not ready for download yet.
12. If the report cannot be downloaded for another reason, return status "blocked" or "partial" with current_url, visible_screen, blocked_by, and whether report creation was attempted.
13. If a report is downloaded, return status "partial" with source "downloaded_ad_report_pending_parse"; the app will parse the downloaded CSV/Excel after the session.
14. In pickup mode, always end with one of these results: downloaded report, blocked because report is still processing, or blocked because matching report was not found. Do not continue exploring beyond that.
15. Return only the requested structured result. Use null for metrics that are not visible in the browser; downloaded file parsing will fill exact values.
"""


def _build_amazon_ads_report_pickup_task(client: dict[str, Any], question: str) -> str:
    now = datetime.now(ZoneInfo("Asia/Tokyo"))
    today = now.date().isoformat()
    date_range_instruction = _date_range_instruction(question, now)
    company_name = client.get("name") or ""
    shop_name = client.get("shop_name") or ""
    marketplace = client.get("marketplace") or "Amazon.co.jp"
    return f"""
You are controlling an already-authenticated browser profile for Amazon Ads Japan.

Safety rules:
- This is a short pickup-only run. You may download an already-created Amazon Ads report, but you must not create a new report in this task.
- Do not type, request, reveal, or store passwords, 2FA codes, API keys, recovery codes, signed URLs, or other secrets.
- If Amazon asks for login, password, passkey, CAPTCHA, or 2FA, stop immediately and return status "needs_user_login".
- Do not change bids, budgets, campaigns, listings, account settings, billing settings, or payments.
- Do not click any button that confirms a purchase, subscription, payment, campaign launch, paid upgrade, billing change, budget change, or advertising change.
- If a paid confirmation, budget confirmation, billing confirmation, or advertising-change confirmation appears, stop immediately and return status "needs_confirmation".

Client:
- Company name: {company_name}
- Seller Central shop/store name to select: {shop_name}
- Marketplace: {marketplace}
- Today in Japan: {today}
- User question: {question}
- Date instruction: {date_range_instruction}

Task:
1. Open the Amazon Ads console directly: https://advertising.amazon.co.jp/
2. Confirm that the active advertiser/store/account matches "{shop_name}" or visibly shows "{shop_name}". If a selector is visible and an exact or very close match exists, choose it.
3. Navigate only to Sponsored ads reports / 広告レポート / レポート. Use the visible Reports navigation if a direct reports page is not already available.
4. Open only the One-time / 1回限り reports tab. Do not open the report creation form.
5. Refresh the browser page once, then reopen the One-time / 1回限り tab if needed.
6. Check only the visible report rows and at most one short scroll within the report table. Do not run broad JavaScript/page-wide searches, do not keep scanning the whole page, and do not paginate through many pages.
7. Find a row matching the requested period. For April 2026, match 2026-04-01 through 2026-04-30. Prefer Sponsored Products / スポンサープロダクト and campaign / キャンペーン scope.
8. If the matching row is complete or shows a download button/icon/menu item, click the download control exactly once and then stop.
9. If the matching row is visible but pending, processing, queued, or not ready, return status "blocked" with blocked_by "report_processing_not_ready".
10. If no matching row is confidently visible after the single refresh and short table check, return status "blocked" with blocked_by "matching_report_not_found".
11. Return status "partial" with source "downloaded_ad_report_pending_parse" if you clicked a download control. The app will parse the downloaded CSV/Excel after the session.
12. Return only the requested structured result. Use null for metrics that are not visible in the browser; downloaded file parsing will fill exact values.
"""


def _build_seller_central_access_check_task(client: dict[str, Any], question: str) -> str:
    today = datetime.now(ZoneInfo("Asia/Tokyo")).date().isoformat()
    company_name = client.get("name") or ""
    shop_name = client.get("shop_name") or ""
    marketplace = client.get("marketplace") or "Amazon.co.jp"
    return f"""
You are controlling an already-authenticated browser profile for Amazon Seller Central Japan.

Safety rules:
- Read-only navigation only.
- Do not fetch reports, download files, change settings, change ads, change budgets, or edit listings.
- Do not type, request, reveal, or store passwords, 2FA codes, API keys, recovery codes, or other secrets.
- If Amazon asks for login, password, passkey, CAPTCHA, or 2FA, stop immediately and return the current screen state.
- Do not click any button that confirms a purchase, subscription, payment, paid upgrade, campaign launch, or billing change.
- If a paid confirmation or billing confirmation appears, stop immediately and return the current screen state.

Client:
- Company name: {company_name}
- Seller Central shop/store name to select: {shop_name}
- Marketplace: {marketplace}
- Today in Japan: {today}
- User question: {question}

Task:
1. Open https://sellercentral.amazon.co.jp/
2. Identify whether Seller Central is visible, whether login/2FA is required, and whether the selected shop/store is "{shop_name}".
3. If Seller Central home is visible, do not wait on the home page. Continue to the advertising access check.
4. If a store selector is visible and "{shop_name}" is available, choose it. Do not continue if this requires login, 2FA, payment confirmation, or a paid upgrade confirmation.
5. Try to reach an advertising or campaign manager screen using direct navigation first:
   - https://advertising.amazon.co.jp/cm/campaigns
   - https://advertising.amazon.co.jp/
   If direct navigation redirects back to Seller Central, use the visible Seller Central navigation/search to open "広告", "キャンペーンマネージャー", "広告キャンペーン", or "Advertising".
6. Stop as soon as an Amazon Ads, advertising, campaigns, or campaign manager page is visible. Do not retrieve advertising metrics.
7. If you are still on Seller Central home after trying the direct advertising URL and visible navigation once, return status "blocked" with blocked_by "seller_central_home_navigation".
8. Return the requested structured result with the current URL, page title, visible screen description, and blocker if any.
"""


def _result_from_session(session: dict[str, Any], summary: str | None = None) -> BrowserUseRunResult:
    status = str(session.get("status") or "unknown")
    output = session.get("output")
    if isinstance(output, str):
        try:
            output = json.loads(output)
        except json.JSONDecodeError:
            pass
    output = _enrich_estimated_metrics(output)
    output_summary = _output_summary(output)
    if not output_summary:
        output_summary = summary or str(session.get("lastStepSummary") or "")
    if not output_summary:
        output_summary = "browser-useの実行結果を取得しました。"
    is_success = session.get("isTaskSuccessful")
    if is_success is None and _output_is_success(output):
        is_success = True
    return BrowserUseRunResult(
        attempted=True,
        status=status,
        summary=output_summary,
        session_id=session.get("id"),
        live_url=session.get("liveUrl"),
        last_step_summary=session.get("lastStepSummary"),
        output=output,
        is_success=is_success,
        total_cost_usd=str(session.get("totalCostUsd")) if session.get("totalCostUsd") is not None else None,
        screenshot_url=session.get("screenshotUrl"),
    )


def run_amazon_ads_report_pickup(client: dict[str, Any], question: str) -> BrowserUseRunResult:
    config = get_browser_use_config()
    if not config.api_key_configured:
        return BrowserUseRunResult(
            attempted=False,
            status="unavailable",
            summary="BROWSER_USE_API_KEYが未設定のため、広告レポート確認は実行していません。",
        )

    account_key = client.get("seller_account_key")
    profile_id = seller_central_profile_id_for_account(account_key)
    if not profile_id:
        return BrowserUseRunResult(
            attempted=False,
            status="unavailable",
            summary="選択中クライアントに対応するbrowser-useプロフィールIDが未設定です。",
        )
    if not client.get("shop_name"):
        return BrowserUseRunResult(
            attempted=False,
            status="unavailable",
            summary="選択中クライアントのショップ名が未設定です。",
        )

    timeout_raw = _secret_or_env("BROWSER_USE_REPORT_PICKUP_TIMEOUT_SECONDS", "150") or "150"
    try:
        pickup_timeout_seconds = max(60, min(240, int(timeout_raw)))
    except ValueError:
        pickup_timeout_seconds = 150

    payload: dict[str, Any] = {
        "task": _build_amazon_ads_report_pickup_task(client, question),
        "model": _normalize_v3_model(config.default_model),
        "keepAlive": False,
        "maxCostUsd": _secret_or_env("BROWSER_USE_REPORT_PICKUP_MAX_COST_USD", "0.45") or "0.45",
        "profileId": profile_id,
        "proxyCountryCode": config.proxy_country_code,
        "outputSchema": SELLER_CENTRAL_OUTPUT_SCHEMA,
        "enableRecording": False,
        "skills": True,
        "agentmail": False,
        "codeMode": False,
        "cacheScript": False,
        "autoHeal": True,
    }

    try:
        session = _browser_use_request("POST", "/api/v3/sessions", payload)
        session_id = session.get("id")
        if not session_id:
            return BrowserUseRunResult(
                attempted=True,
                status="failed",
                summary="browser-useセッションIDを取得できませんでした。",
                output=session,
            )

        final_statuses = {"stopped", "timed_out", "error"}
        deadline = time.monotonic() + pickup_timeout_seconds
        while str(session.get("status") or "") not in final_statuses and time.monotonic() < deadline:
            time.sleep(4)
            session = _browser_use_request("GET", f"/api/v3/sessions/{session_id}")

        if str(session.get("status") or "") not in final_statuses:
            latest = get_browser_use_session(session_id)
            stop_result = stop_browser_use_session(session_id)
            if stop_result.status != "error":
                output = stop_result.output or latest.output
                output_has_result = _output_has_result(output)
                output_success = _output_is_success(output)
                result = BrowserUseRunResult(
                    attempted=True,
                    status=stop_result.status,
                    summary=(
                        f"{_output_summary(output)} セッションは短時間確認の上限で自動停止しました。"
                        if output_has_result
                        else "短時間のレポート一覧確認が完了しなかったため、コスト抑制のためbrowser-useセッションを自動停止しました。"
                    ),
                    session_id=stop_result.session_id or session_id,
                    live_url=stop_result.live_url or latest.live_url or session.get("liveUrl"),
                    last_step_summary=stop_result.last_step_summary or latest.last_step_summary or session.get("lastStepSummary"),
                    output=output,
                    is_success=True if output_success else stop_result.is_success,
                    total_cost_usd=stop_result.total_cost_usd or latest.total_cost_usd,
                    screenshot_url=stop_result.screenshot_url or latest.screenshot_url,
                )
                return _merge_report_downloads(result)
            result = BrowserUseRunResult(
                attempted=True,
                status=str(session.get("status") or "running"),
                summary="browser-useはまだ実行中です。自動停止も失敗したため、ライブ画面から手動停止してください。",
                session_id=session_id,
                live_url=session.get("liveUrl"),
                last_step_summary=session.get("lastStepSummary"),
                output=session.get("output"),
                is_success=session.get("isTaskSuccessful"),
                total_cost_usd=str(session.get("totalCostUsd")) if session.get("totalCostUsd") is not None else None,
                screenshot_url=session.get("screenshotUrl"),
                error=stop_result.error,
            )
            return _merge_report_downloads(result)
        return _merge_report_downloads(_result_from_session(session))
    except BrowserUseError as exc:
        return BrowserUseRunResult(
            attempted=True,
            status="error",
            summary="browser-use API呼び出しに失敗しました。",
            error=str(exc),
        )


def run_seller_central_metrics_fetch(client: dict[str, Any], question: str) -> BrowserUseRunResult:
    config = get_browser_use_config()
    if not config.api_key_configured:
        return BrowserUseRunResult(
            attempted=False,
            status="unavailable",
            summary="BROWSER_USE_API_KEYが未設定のため、Seller Central自動取得は実行していません。",
        )

    account_key = client.get("seller_account_key")
    profile_id = seller_central_profile_id_for_account(account_key)
    if not profile_id:
        return BrowserUseRunResult(
            attempted=False,
            status="unavailable",
            summary="選択中クライアントに対応するbrowser-useプロフィールIDが未設定です。",
        )
    if not client.get("shop_name"):
        return BrowserUseRunResult(
            attempted=False,
            status="unavailable",
            summary="選択中クライアントのショップ名が未設定です。",
        )

    payload: dict[str, Any] = {
        "task": _build_seller_central_task(client, question),
        "model": _normalize_v3_model(config.default_model),
        "keepAlive": False,
        "maxCostUsd": (
            _secret_or_env("BROWSER_USE_REPORT_PICKUP_MAX_COST_USD", "0.60")
            if _is_report_pickup_question(question)
            else _secret_or_env("BROWSER_USE_REPORT_MAX_COST_USD", "0.85")
        )
        or "0.85",
        "profileId": profile_id,
        "proxyCountryCode": config.proxy_country_code,
        "outputSchema": SELLER_CENTRAL_OUTPUT_SCHEMA,
        "enableRecording": False,
        "skills": True,
        "agentmail": False,
        "codeMode": False,
        "cacheScript": False,
        "autoHeal": True,
    }

    try:
        session = _browser_use_request("POST", "/api/v3/sessions", payload)
        session_id = session.get("id")
        if not session_id:
            return BrowserUseRunResult(
                attempted=True,
                status="failed",
                summary="browser-useセッションIDを取得できませんでした。",
                output=session,
            )

        final_statuses = {"stopped", "timed_out", "error"}
        deadline = time.monotonic() + (
            max(config.poll_timeout_seconds, 300)
            if _is_report_pickup_question(question)
            else max(config.poll_timeout_seconds, 420)
        )
        while str(session.get("status") or "") not in final_statuses and time.monotonic() < deadline:
            time.sleep(4)
            session = _browser_use_request("GET", f"/api/v3/sessions/{session_id}")

        if str(session.get("status") or "") not in final_statuses:
            latest = get_browser_use_session(session_id)
            stop_result = stop_browser_use_session(session_id)
            if stop_result.status != "error":
                output = stop_result.output or latest.output
                output_has_result = _output_has_result(output)
                output_success = _output_is_success(output)
                result = BrowserUseRunResult(
                    attempted=True,
                    status=stop_result.status,
                    summary=(
                        f"{_output_summary(output)} セッションはコスト抑制のため自動停止しました。"
                        if output_has_result
                        else "規定時間内に完了しなかったため、コスト抑制のためbrowser-useセッションを自動停止しました。"
                    ),
                    session_id=stop_result.session_id or session_id,
                    live_url=stop_result.live_url or latest.live_url or session.get("liveUrl"),
                    last_step_summary=stop_result.last_step_summary or latest.last_step_summary or session.get("lastStepSummary"),
                    output=output,
                    is_success=True if output_success else stop_result.is_success,
                    total_cost_usd=stop_result.total_cost_usd or latest.total_cost_usd,
                    screenshot_url=stop_result.screenshot_url or latest.screenshot_url,
                )
                return _merge_report_downloads(result)
            result = BrowserUseRunResult(
                attempted=True,
                status=str(session.get("status") or "running"),
                summary="browser-useはまだ実行中です。自動停止も失敗したため、ライブ画面から手動停止してください。",
                session_id=session_id,
                live_url=session.get("liveUrl"),
                last_step_summary=session.get("lastStepSummary"),
                output=session.get("output"),
                is_success=session.get("isTaskSuccessful"),
                total_cost_usd=str(session.get("totalCostUsd")) if session.get("totalCostUsd") is not None else None,
                screenshot_url=session.get("screenshotUrl"),
                error=stop_result.error,
            )
            return _merge_report_downloads(result)
        return _merge_report_downloads(_result_from_session(session))
    except BrowserUseError as exc:
        return BrowserUseRunResult(
            attempted=True,
            status="error",
            summary="browser-use API呼び出しに失敗しました。",
            error=str(exc),
        )


def run_seller_central_access_check(client: dict[str, Any], question: str) -> BrowserUseRunResult:
    config = get_browser_use_config()
    if not config.api_key_configured:
        return BrowserUseRunResult(
            attempted=False,
            status="unavailable",
            summary="BROWSER_USE_API_KEYが未設定のため、Seller Central到達確認は実行していません。",
        )

    account_key = client.get("seller_account_key")
    profile_id = seller_central_profile_id_for_account(account_key)
    if not profile_id:
        return BrowserUseRunResult(
            attempted=False,
            status="unavailable",
            summary="選択中クライアントに対応するbrowser-useプロフィールIDが未設定です。",
        )
    if not client.get("shop_name"):
        return BrowserUseRunResult(
            attempted=False,
            status="unavailable",
            summary="選択中クライアントのショップ名が未設定です。",
        )

    payload: dict[str, Any] = {
        "task": _build_seller_central_access_check_task(client, question),
        "model": _normalize_v3_model(config.default_model),
        "keepAlive": False,
        "maxCostUsd": config.max_cost_usd,
        "profileId": profile_id,
        "proxyCountryCode": config.proxy_country_code,
        "outputSchema": SELLER_CENTRAL_ACCESS_CHECK_SCHEMA,
        "enableRecording": False,
        "skills": True,
        "agentmail": False,
        "codeMode": False,
        "cacheScript": False,
        "autoHeal": True,
    }

    try:
        session = _browser_use_request("POST", "/api/v3/sessions", payload)
        session_id = session.get("id")
        if not session_id:
            return BrowserUseRunResult(
                attempted=True,
                status="failed",
                summary="browser-useセッションIDを取得できませんでした。",
                output=session,
            )

        final_statuses = {"stopped", "timed_out", "error"}
        deadline = time.monotonic() + config.poll_timeout_seconds
        while str(session.get("status") or "") not in final_statuses and time.monotonic() < deadline:
            time.sleep(4)
            session = _browser_use_request("GET", f"/api/v3/sessions/{session_id}")

        if str(session.get("status") or "") not in final_statuses:
            latest = get_browser_use_session(session_id)
            stop_result = stop_browser_use_session(session_id)
            if stop_result.status != "error":
                output = stop_result.output or latest.output
                output_has_result = _output_has_result(output)
                output_success = _output_is_success(output)
                return BrowserUseRunResult(
                    attempted=True,
                    status=stop_result.status,
                    summary=(
                        f"{_output_summary(output)} セッションはコスト抑制のため自動停止しました。"
                        if output_has_result
                        else "規定時間内に到達確認が完了しなかったため、コスト抑制のためbrowser-useセッションを自動停止しました。"
                    ),
                    session_id=stop_result.session_id or session_id,
                    live_url=stop_result.live_url or latest.live_url or session.get("liveUrl"),
                    last_step_summary=stop_result.last_step_summary or latest.last_step_summary or session.get("lastStepSummary"),
                    output=output,
                    is_success=True if output_success else stop_result.is_success,
                    total_cost_usd=stop_result.total_cost_usd or latest.total_cost_usd,
                    screenshot_url=stop_result.screenshot_url or latest.screenshot_url,
                )
            return BrowserUseRunResult(
                attempted=True,
                status=str(session.get("status") or "running"),
                summary="browser-useはまだ実行中です。自動停止も失敗したため、ライブ画面から手動停止してください。",
                session_id=session_id,
                live_url=session.get("liveUrl"),
                last_step_summary=session.get("lastStepSummary"),
                output=session.get("output"),
                is_success=session.get("isTaskSuccessful"),
                total_cost_usd=str(session.get("totalCostUsd")) if session.get("totalCostUsd") is not None else None,
                screenshot_url=session.get("screenshotUrl"),
                error=stop_result.error,
            )
        return _result_from_session(session)
    except BrowserUseError as exc:
        return BrowserUseRunResult(
            attempted=True,
            status="error",
            summary="browser-use API呼び出しに失敗しました。",
            error=str(exc),
        )
